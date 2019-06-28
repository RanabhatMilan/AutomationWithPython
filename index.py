import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path

def process_spreedsheets(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    n_row = sheet.max_row+1
    for row in range(2, n_row):
        cell = sheet.cell(row, 6)
        corrected_value = cell.value * 100
        new_row = sheet.cell(row, 9)
        new_row.value = corrected_value

    chart = BarChart()
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=6, max_col=6)
    chart.add_data(values)
    sheet.add_chart(chart, "A"+str(n_row))
    wb.save(filename)

path = Path("Spreedsheets")
if(path.exists()):
    for file in path.glob('*'):
       process_spreedsheets(file)









